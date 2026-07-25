"""ExcelExportService 测试。

验证目标（设计文档第 12.2 节）：
1. generate_excel 返回文件路径
2. 文件可被 openpyxl 重新打开
3. 含 4 个 sheet：分发记录/收录检测/AI 采信/数据汇总
4. 表头正确
"""
import os
import pytest
from openpyxl import load_workbook

from app.services.excel_export_service import ExcelExportService


@pytest.mark.asyncio
async def test_generate_excel_creates_file(tmp_path):
    """generate_excel 生成 Excel 文件。"""
    service = ExcelExportService(output_dir=str(tmp_path))

    export_data = {
        "client_name": "测试客户",
        "date_from": "2026-07-01",
        "date_to": "2026-07-25",
        "distributions": [
            {
                "remote_url": "https://example.com/1",
                "content_title": "文章1",
                "source": "geoflow",
                "index_status": {"baidu": "indexed", "bing": "pending"},
            }
        ],
        "index_results": [
            {"url": "https://example.com/1", "baidu": "indexed", "bing": "pending"}
        ],
        "citation_results": [
            {"url": "https://example.com/1", "model": "qwen", "hit_type": "direct"}
        ],
        "summary": {
            "total_distributions": 1,
            "indexed_count": 1,
            "citation_count": 1,
        },
    }

    file_path = await service.generate_excel(export_data, filename="test.xlsx")

    assert os.path.exists(file_path)
    assert file_path.endswith(".xlsx")


@pytest.mark.asyncio
async def test_excel_has_4_sheets(tmp_path):
    """Excel 含 4 个 sheet。"""
    service = ExcelExportService(output_dir=str(tmp_path))
    export_data = {
        "distributions": [], "index_results": [], "citation_results": [], "summary": {},
    }
    file_path = await service.generate_excel(export_data, filename="test_sheets.xlsx")

    wb = load_workbook(file_path)
    assert "分发记录" in wb.sheetnames
    assert "收录检测" in wb.sheetnames
    assert "AI采信" in wb.sheetnames
    assert "数据汇总" in wb.sheetnames


@pytest.mark.asyncio
async def test_excel_distribution_sheet_headers(tmp_path):
    """分发记录 sheet 表头正确。"""
    service = ExcelExportService(output_dir=str(tmp_path))
    file_path = await service.generate_excel(
        {"distributions": [], "index_results": [], "citation_results": [], "summary": {}},
        filename="test_headers.xlsx",
    )
    wb = load_workbook(file_path)
    ws = wb["分发记录"]
    headers = [cell.value for cell in ws[1]]
    assert "序号" in headers
    assert "文章标题" in headers
    assert "URL" in headers
    assert "来源" in headers
