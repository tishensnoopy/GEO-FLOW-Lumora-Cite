"""Excel 明细导出服务——openpyxl 4 sheet 工作簿。

设计文档第 12.2 节。

Sheet 结构：
1. 分发记录：序号/文章标题/URL/来源/分发时间/状态
2. 收录检测：URL/百度/头条/搜狗/360/必应/检测时间
3. AI采信：URL/模型/问题/命中类型/检测时间
4. 数据汇总：统计指标
"""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ExcelExportService:
    def __init__(self, output_dir: str = "/app/exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    async def generate_excel(
        self,
        export_data: dict,
        filename: str | None = None,
    ) -> str:
        """生成 Excel 报告。"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{timestamp}.xlsx"

        file_path = str(self.output_dir / filename)

        wb = Workbook()
        # 默认 sheet 重命名为"分发记录"
        ws_dist = wb.active
        ws_dist.title = "分发记录"
        self._fill_distribution_sheet(ws_dist, export_data.get("distributions", []))

        ws_index = wb.create_sheet("收录检测")
        self._fill_index_sheet(ws_index, export_data.get("index_results", []))

        ws_citation = wb.create_sheet("AI采信")
        self._fill_citation_sheet(ws_citation, export_data.get("citation_results", []))

        ws_summary = wb.create_sheet("数据汇总")
        self._fill_summary_sheet(ws_summary, export_data.get("summary", {}))

        wb.save(file_path)
        return file_path

    def _style_header(self, ws, headers: list[str]):
        """设置表头样式。"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def _fill_distribution_sheet(self, ws, distributions: list[dict]):
        headers = ["序号", "文章标题", "URL", "来源", "分发时间", "状态"]
        self._style_header(ws, headers)
        for idx, dist in enumerate(distributions, 1):
            ws.cell(row=idx + 1, column=1, value=idx)
            ws.cell(row=idx + 1, column=2, value=dist.get("content_title", ""))
            ws.cell(row=idx + 1, column=3, value=dist.get("remote_url", ""))
            ws.cell(row=idx + 1, column=4, value=dist.get("source", ""))
            ws.cell(row=idx + 1, column=5, value=dist.get("distributed_at", ""))
            ws.cell(row=idx + 1, column=6, value=dist.get("status", ""))

    def _fill_index_sheet(self, ws, index_results: list[dict]):
        headers = ["URL", "百度", "头条", "搜狗", "360", "必应", "检测时间"]
        self._style_header(ws, headers)
        for idx, ir in enumerate(index_results, 1):
            ws.cell(row=idx + 1, column=1, value=ir.get("url", ""))
            ws.cell(row=idx + 1, column=2, value=ir.get("baidu", ir.get("baidu_status", "")))
            ws.cell(row=idx + 1, column=3, value=ir.get("toutiao", ir.get("toutiao_status", "")))
            ws.cell(row=idx + 1, column=4, value=ir.get("sogou", ir.get("sogou_status", "")))
            ws.cell(row=idx + 1, column=5, value=ir.get("so360", ir.get("so360_status", "")))
            ws.cell(row=idx + 1, column=6, value=ir.get("bing", ir.get("bing_status", "")))
            ws.cell(row=idx + 1, column=7, value=ir.get("checked_at", ""))

    def _fill_citation_sheet(self, ws, citations: list[dict]):
        headers = ["URL", "模型", "问题", "命中类型", "检测时间"]
        self._style_header(ws, headers)
        for idx, c in enumerate(citations, 1):
            ws.cell(row=idx + 1, column=1, value=c.get("url", ""))
            ws.cell(row=idx + 1, column=2, value=c.get("model", ""))
            ws.cell(row=idx + 1, column=3, value=c.get("question", ""))
            ws.cell(row=idx + 1, column=4, value=c.get("hit_type", ""))
            ws.cell(row=idx + 1, column=5, value=c.get("checked_at", ""))

    def _fill_summary_sheet(self, ws, summary: dict):
        headers = ["指标", "数值"]
        self._style_header(ws, headers)
        rows = [
            ("分发总数", summary.get("total_distributions", 0)),
            ("已收录数", summary.get("indexed_count", 0)),
            ("AI 采信数", summary.get("citation_count", 0)),
            ("平均收录率", summary.get("avg_index_rate", 0)),
        ]
        for idx, (label, value) in enumerate(rows, 2):
            ws.cell(row=idx, column=1, value=label)
            ws.cell(row=idx, column=2, value=value)
